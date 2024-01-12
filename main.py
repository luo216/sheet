import yaml
from openpyxl import load_workbook


def update():
    # 读取YAML文件
    with open("./setup.yaml", "r") as file:
        yaml_data = yaml.safe_load(file)

    # 获取Excel文件路径
    excel_path = yaml_data["excel_path"]
    workbook = load_workbook(excel_path)

    # 遍历关键字列表
    keyword = yaml_data["keywords"]

    def load_data(i):
        sheet_name = keyword[i]["sheet_name"]
        sheet = workbook[sheet_name]
        start_row = keyword[i]["start_add"]["row"]
        start_col = keyword[i]["start_add"]["col"]
        length = keyword[i]["length"]

        arr = []
        for i in range(length):
            arr.append(sheet.cell(row=start_row + i, column=start_col).value)

        # 去除空值
        arr = [x for x in arr if x is not None]
        return arr

    # 定义一个字典用于存储所有数据
    data = {}

    for i in range(len(keyword)):
        # data中存储字典，key为关键字,value为数据
        data[keyword[i]["name"]] = load_data(i)

    return data


def main():
    data = update()
    print(data)


if __name__ == "__main__":
    main()
